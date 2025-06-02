import os
import pandas as pd
from flask import Flask, jsonify, request, send_file, render_template
from flask_cors import CORS
import datetime
import numpy as np
import sqlite3
from io import BytesIO
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# --- Configuración de Archivos y Columnas ---
APP_ROOT = os.path.dirname(os.path.abspath(__file__))
DATABASE_FOLDER = os.path.join(APP_ROOT, 'databases')
ITEM_MASTER_CSV_PATH = os.path.join(DATABASE_FOLDER, 'AURRSGLBD0250 - Item Stockroom Balance.csv')
GRN_CSV_FILE_PATH = os.path.join(DATABASE_FOLDER, 'AURRSGLBD0280 - Stock In Goods Inwards And Inspection.csv')
DB_FILE_PATH = os.path.join(APP_ROOT, 'inbound_log.db')

COLUMNS_TO_READ_MASTER = [
    'Item_Code', 'Item_Description', 'Weight_per_Unit',
    'Bin_1', 'Aditional_Bin_Location',
]
GRN_COLUMN_NAME_IN_CSV = 'GRN_Number' 
COLUMNS_TO_READ_GRN = [GRN_COLUMN_NAME_IN_CSV, 'Item_Code', 'Quantity']

app = Flask(__name__, template_folder='templates', static_folder='static')
CORS(app)

# --- Funciones de Manejo de CSV ---
def read_csv_safe(file_path, columns=None):
    if not os.path.exists(file_path):
        print(f"Error CSV: Archivo no encontrado en {file_path}")
        return None
    try:
        df = pd.read_csv(file_path, usecols=columns, dtype=str, keep_default_na=True)
        df = df.replace({np.nan: None})
        return df
    except Exception as e:
        print(f"Error CSV: Error inesperado leyendo CSV {file_path}: {e}")
        return None

def get_item_details_from_master_csv(item_code):
    df_master = read_csv_safe(ITEM_MASTER_CSV_PATH, columns=COLUMNS_TO_READ_MASTER)
    if df_master is None: return None
    result = df_master[df_master['Item_Code'] == item_code]
    return result.iloc[0].fillna('').to_dict() if not result.empty else None

def get_grn_specific_expected_quantity(import_ref_form, item_code_form):
    expected_quantity = 0
    df_grn = read_csv_safe(GRN_CSV_FILE_PATH, columns=COLUMNS_TO_READ_GRN)
    if df_grn is None:
        print(f"Advertencia CSV: No se pudo leer el archivo GRN {GRN_CSV_FILE_PATH}.")
        return expected_quantity
    result = df_grn[
        (df_grn[GRN_COLUMN_NAME_IN_CSV] == import_ref_form) &
        (df_grn['Item_Code'] == item_code_form)
    ]
    if not result.empty:
        quantity_value = result.iloc[0].get('Quantity')
        if quantity_value is not None:
            try:
                numeric_value = pd.to_numeric(quantity_value, errors='coerce')
                if pd.notna(numeric_value): expected_quantity = int(numeric_value)
                else: print(f"Advertencia CSV: Cantidad no numérica ('{quantity_value}') para GRN {import_ref_form}, Item {item_code_form}.")
            except (ValueError, TypeError): print(f"Advertencia CSV: Error convirtiendo cantidad ('{quantity_value}') para GRN {import_ref_form}, Item {item_code_form}.")
        else: print(f"Advertencia CSV: Columna 'Quantity' vacía para GRN {import_ref_form}, Item {item_code_form}.")
    return expected_quantity

# --- Funciones de Base de Datos SQLite ---
def init_db():
    try:
        conn = sqlite3.connect(DB_FILE_PATH)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT NOT NULL, importRef TEXT,
                waybill TEXT, itemCode TEXT, itemDescription TEXT, binLocation TEXT,
                relocatedBin TEXT, qtyReceived INTEGER, qtyGrn INTEGER, difference INTEGER
            )''') 
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_importRef_itemCode ON logs (importRef, itemCode)")
        conn.commit()
        print(f"DB: Base de datos SQLite inicializada/verificada en: {DB_FILE_PATH}")
    except sqlite3.Error as e: print(f"DB Error (init_db): {e}")
    finally:
        if conn: conn.close()

def save_log_entry_db(entry_data):
    conn = None
    try:
        conn = sqlite3.connect(DB_FILE_PATH, timeout=10)
        cursor = conn.cursor()
        sql = '''INSERT INTO logs (timestamp, importRef, waybill, itemCode, itemDescription,
                              binLocation, relocatedBin, qtyReceived, qtyGrn, difference)
                 VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''
        values = (
            entry_data.get('timestamp'), entry_data.get('importRef'), entry_data.get('waybill'),
            entry_data.get('itemCode'), entry_data.get('itemDescription'), entry_data.get('binLocation'),
            entry_data.get('relocateBin'), 
            entry_data.get('qtyReceived'), entry_data.get('qtyGrn'),
            entry_data.get('difference')
        )
        cursor.execute(sql, values)
        entry_id = cursor.lastrowid
        conn.commit()
        print(f"DB (save_log_entry_db): Entrada guardada con ID: {entry_id}")
        return entry_id
    except sqlite3.Error as e:
        print(f"DB Error (save_log_entry_db): {e}")
        if conn: conn.rollback()
        return None
    finally:
        if conn: conn.close()

def update_log_entry_db(log_id, entry_data_for_db): 
    conn = None
    try:
        conn = sqlite3.connect(DB_FILE_PATH, timeout=10)
        cursor = conn.cursor()
        sql = '''UPDATE logs SET 
                    waybill = ?, 
                    relocatedBin = ?, 
                    qtyReceived = ?, 
                    difference = ?,
                    timestamp = ? 
                 WHERE id = ?'''
        values = (
            entry_data_for_db.get('waybill'), 
            entry_data_for_db.get('relocatedBin'), # Clave de BD 'relocatedBin' (con d)
            entry_data_for_db.get('qtyReceived'), 
            entry_data_for_db.get('difference'),
            entry_data_for_db.get('timestamp'), 
            log_id
        )
        cursor.execute(sql, values)
        conn.commit()
        print(f"DB (update_log_entry_db): Filas actualizadas para ID {log_id}: {cursor.rowcount}")
        return cursor.rowcount > 0
    except sqlite3.Error as e:
        print(f"DB Error (update_log_entry_db) para ID {log_id}: {e}")
        if conn: conn.rollback()
        return False
    finally:
        if conn: conn.close()

def load_log_data_db():
    logs = []
    conn = None
    try:
        conn = sqlite3.connect(DB_FILE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM logs ORDER BY id DESC")
        rows = cursor.fetchall()
        for row in rows: logs.append(dict(row)) 
        print(f"DB (load_log_data_db): Cargados {len(logs)} registros.") # DEBUG
        return logs
    except sqlite3.Error as e:
        print(f"DB Error (load_log_data_db): {e}")
        return []
    finally:
        if conn: conn.close()

def get_log_entry_by_id(log_id):
    conn = None
    try:
        conn = sqlite3.connect(DB_FILE_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM logs WHERE id = ?", (log_id,))
        row = cursor.fetchone()
        return dict(row) if row else None
    except sqlite3.Error as e:
        print(f"DB Error (get_log_entry_by_id) para ID {log_id}: {e}")
        return None
    finally:
        if conn: conn.close()

def get_total_received_for_grn_item(import_ref, item_code):
    total_received = 0
    conn = None
    try:
        conn = sqlite3.connect(DB_FILE_PATH)
        cursor = conn.cursor()
        sql = "SELECT SUM(qtyReceived) FROM logs WHERE importRef = ? AND itemCode = ?"
        cursor.execute(sql, (import_ref, item_code))
        result = cursor.fetchone()
        if result and result[0] is not None: total_received = int(result[0])
    except sqlite3.Error as e: print(f"DB Error (get_total_received_for_grn_item): {e}")
    finally:
        if conn: conn.close()
    return total_received

# --- Endpoints de la API ---
@app.route('/api/find_item/<item_code>/<import_ref>', methods=['GET'])
def find_item(item_code, import_ref):
    item_details = get_item_details_from_master_csv(item_code)
    if item_details is None:
        return jsonify({"error": f"Artículo {item_code} no encontrado en el maestro."}), 404
    expected_quantity = get_grn_specific_expected_quantity(import_ref, item_code)
    response_data = {
        "itemCode": item_details.get('Item_Code', item_code),
        "description": item_details.get('Item_Description', 'N/A'),
        "binLocation": item_details.get('Bin_1', 'N/A'),
        "aditionalBins": item_details.get('Aditional_Bin_Location', 'N/A'),
        "weight": item_details.get('Weight_per_Unit', 'N/A'),
        "defaultQtyGrn": expected_quantity
    }
    return jsonify(response_data), 200

@app.route('/api/add_log', methods=['POST'])
def add_log():
    data = request.get_json()
    print(f"API (add_log): Recibido para log: {data}")
    required_fields = ['importRef', 'waybill', 'itemCode', 'quantity']
    if not data or not all(field in data for field in required_fields):
        return jsonify({"error": "Faltan datos requeridos"}), 400
    try:
        quantity_received_now = int(data['quantity'])
        if quantity_received_now <= 0: return jsonify({"error": "Cantidad debe ser > 0"}), 400
    except (ValueError, TypeError): return jsonify({"error": "Cantidad debe ser número"}), 400

    item_code = data['itemCode']
    import_ref = data['importRef']
    item_details = get_item_details_from_master_csv(item_code)
    if item_details is None: return jsonify({"error": f"Artículo {item_code} no existe."}), 400
    
    item_description = item_details.get('Item_Description', 'N/A')
    original_bin = item_details.get('Bin_1', 'N/A')
    expected_quantity_grn = get_grn_specific_expected_quantity(import_ref, item_code)
    
    total_already_received_for_grn_item = get_total_received_for_grn_item(import_ref, item_code)
    new_cumulative_total_received = total_already_received_for_grn_item + quantity_received_now
    cumulative_difference = new_cumulative_total_received - expected_quantity_grn

    log_entry_data_for_db = { 
        "timestamp": datetime.datetime.now().isoformat(timespec='seconds'),
        "importRef": import_ref, 
        "waybill": data['waybill'], 
        "itemCode": item_code,
        "itemDescription": item_description, 
        "binLocation": original_bin,
        "relocatedBin": data.get('relocateBin', ''), 
        "qtyReceived": quantity_received_now,
        "qtyGrn": expected_quantity_grn, 
        "difference": cumulative_difference
    }
    
    new_log_id = save_log_entry_db(log_entry_data_for_db) 
    if new_log_id is not None: 
        log_entry_data_for_response = {
            "id": new_log_id,
            "timestamp": log_entry_data_for_db["timestamp"],
            "importRef": log_entry_data_for_db["importRef"],
            "waybill": log_entry_data_for_db["waybill"],
            "itemCode": log_entry_data_for_db["itemCode"],
            "itemDescription": log_entry_data_for_db["itemDescription"],
            "binLocation": log_entry_data_for_db["binLocation"],
            "relocatedBin": data.get('relocateBin', ''), # <-- CORREGIDO
            "qtyReceived": log_entry_data_for_db["qtyReceived"],
            "qtyGrn": log_entry_data_for_db["qtyGrn"],
            "difference": log_entry_data_for_db["difference"]
        }
        print(f"API (add_log): Devolviendo entrada: {log_entry_data_for_response}")
        return jsonify({"message": "Registro añadido con éxito", "entry": log_entry_data_for_response}), 201
    else:
        print(f"API (add_log): Error al guardar, new_log_id es None.")
        return jsonify({"error": "Error interno al guardar registro"}), 500

@app.route('/api/update_log/<int:log_id>', methods=['PUT'])
def update_log(log_id):
    data = request.get_json(silent=True) 
    print(f"API (update_log): Recibido para actualizar log ID {log_id}. Datos parseados: {data}")

    if data is None: 
        return jsonify({"error": "No se recibieron datos JSON válidos."}), 400
    if not isinstance(data, dict):
        return jsonify({"error": "Formato de datos incorrecto."}), 400

    key_from_frontend_waybill = 'waybill'
    key_from_frontend_qtyReceived = 'qtyReceived'
    key_from_frontend_relocateBin = 'relocateBin'

    required_keys_from_frontend = [
        key_from_frontend_waybill,
        key_from_frontend_qtyReceived,
        key_from_frontend_relocateBin 
    ]
    missing_keys = [key for key in required_keys_from_frontend if key not in data]
            
    if missing_keys:
        error_message = f"Faltan datos requeridos para la actualización: {', '.join(missing_keys)}"
        print(f"API (update_log): Validación fallida. {error_message}")
        return jsonify({"error": error_message}), 400
    
    print("API (update_log): Todos los campos requeridos del frontend están presentes.")

    try:
        qty_received_updated = int(data[key_from_frontend_qtyReceived]) 
        if qty_received_updated < 0: 
            return jsonify({"error": "Cantidad recibida no puede ser negativa"}), 400
    except (ValueError, TypeError):
        return jsonify({"error": "Cantidad recibida debe ser un número válido"}), 400

    original_log_entry = get_log_entry_by_id(log_id)
    if not original_log_entry:
        return jsonify({"error": f"Registro de log con ID {log_id} no encontrado."}), 404

    original_qty_grn = original_log_entry['qtyGrn'] 
    updated_difference = qty_received_updated - original_qty_grn

    updated_log_data_for_db = {
        "waybill": data[key_from_frontend_waybill], 
        "relocatedBin": data[key_from_frontend_relocateBin], 
        "qtyReceived": qty_received_updated,
        "difference": updated_difference,
        "timestamp": datetime.datetime.now().isoformat(timespec='seconds'),
    }
    
    if update_log_entry_db(log_id, updated_log_data_for_db):
        full_updated_entry_for_response = {
            "id": log_id,
            "importRef": original_log_entry['importRef'],
            "waybill": updated_log_data_for_db["waybill"],
            "itemCode": original_log_entry['itemCode'],
            "itemDescription": original_log_entry['itemDescription'],
            "binLocation": original_log_entry['binLocation'],
            "relocatedBin": updated_log_data_for_db["relocatedBin"], # <-- CORREGIDO
            "qtyReceived": updated_log_data_for_db["qtyReceived"],
            "qtyGrn": original_log_entry['qtyGrn'],
            "difference": updated_log_data_for_db["difference"],
            "timestamp": updated_log_data_for_db["timestamp"] 
        }
        print(f"API (update_log): Devolviendo entrada actualizada al frontend: {full_updated_entry_for_response}")
        return jsonify({"message": "Registro actualizado con éxito", "entry": full_updated_entry_for_response}), 200
    else:
        return jsonify({"error": "Error interno o ID no encontrado al actualizar el registro en BD"}), 500

@app.route('/api/get_logs', methods=['GET'])
def get_logs():
    logs = load_log_data_db()
    return jsonify(logs), 200

@app.route('/api/export_log', methods=['GET'])
def export_log():
    try:
        logs_data = load_log_data_db()
        print(f"API (export_log): Datos cargados de la BD para exportar: {len(logs_data)} registros.") # DEBUG
        if logs_data and len(logs_data) > 0: # DEBUG: Imprimir una muestra
            print(f"API (export_log): Muestra del primer registro para exportar: {logs_data[0]}")
            if len(logs_data) > 1:
                 print(f"API (export_log): Muestra del último registro para exportar: {logs_data[-1]}")


        if not logs_data:
            return jsonify({"error": "No hay registros para exportar"}), 404
        
        df = pd.DataFrame(logs_data)
        # Asegurarse de que los nombres de columna en la selección coincidan con la BD
        # La columna en la BD es 'relocatedBin' (con 'd')
        df_export = df[[
            'timestamp', 'importRef', 'waybill', 'itemCode', 'itemDescription',
            'binLocation', 'relocatedBin', 'qtyReceived', 'qtyGrn', 'difference'
        ]].rename(columns={
            'timestamp': 'Timestamp', 'importRef': 'GRN 1.', 'waybill': 'Waybill',
            'itemCode': 'Item Code', 'itemDescription': 'Item Description',
            'binLocation': 'Bin Location (Original)', 'relocatedBin': 'Relocated Bin (New)',
            'qtyReceived': 'Qty. Received', 'qtyGrn': 'Qty. GRN', 'difference': 'Difference'
        })
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='InboundLog')
            worksheet = writer.sheets['InboundLog']
            for i, col_name in enumerate(df_export.columns):
                 column_letter = get_column_letter(i + 1)
                 max_len = max(
                     df_export[col_name].astype(str).map(len).max(),
                     len(col_name)
                 ) + 2
                 worksheet.column_dimensions[column_letter].width = max_len
        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inbound_log_{timestamp_str}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error durante la exportación a Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Error interno al generar el archivo Excel"}), 500

@app.route('/api/export_summary', methods=['GET'])
def export_summary():
    # ... (sin cambios significativos, asume que load_log_data_db() es correcto)
    print("API: Solicitud de exportación de resumen (agrupado por GRN) recibida.")
    try:
        logs_data = load_log_data_db()
        if not logs_data:
            print("API Resumen: No hay datos en el log para generar resumen.")
            return jsonify({"error": "No hay registros en el log para generar el resumen"}), 404
        
        df_logs = pd.DataFrame(logs_data)
        required_log_cols = ['importRef', 'itemCode', 'qtyReceived', 'itemDescription', 'qtyGrn']
        if not all(col in df_logs.columns for col in required_log_cols):
             missing_cols = [col for col in required_log_cols if col not in df_logs.columns]
             print(f"API Resumen: Faltan columnas en los logs: {missing_cols}")
             return jsonify({"error": f"Datos de log incompletos para resumen. Faltan: {', '.join(missing_cols)}"}), 500
        
        df_logs['qtyReceived'] = pd.to_numeric(df_logs['qtyReceived'], errors='coerce').fillna(0)
        df_logs['qtyGrn'] = pd.to_numeric(df_logs['qtyGrn'], errors='coerce').fillna(0)

        df_received_summary = df_logs.groupby(['importRef', 'itemCode']).agg(
            totalReceived=('qtyReceived', 'sum'),
            itemDescription=('itemDescription', 'first'),
            totalExpectedGrn=('qtyGrn', 'first') 
        ).reset_index()
        
        df_received_summary['difference'] = df_received_summary['totalReceived'] - df_received_summary['totalExpectedGrn']
        df_received_summary = df_received_summary.sort_values(by=['importRef', 'itemCode'])

        df_final_export = df_received_summary.rename(columns={
            'importRef': 'Número de GRN', 'itemCode': 'Código Ítem',
            'itemDescription': 'Descripción Ítem', 'totalReceived': 'Total Recibido (por GRN)',
            'totalExpectedGrn': 'Total Esperado para Ítem (según Log)', 'difference': 'Diferencia'
        })
        df_final_export = df_final_export[[
            'Número de GRN', 'Código Ítem', 'Descripción Ítem', 
            'Total Recibido (por GRN)', 'Total Esperado para Ítem (según Log)', 'Diferencia'
        ]]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_final_export.to_excel(writer, index=False, sheet_name='ResumenPorGRN')
            worksheet = writer.sheets['ResumenPorGRN']
            blue_font = Font(color="0000FF", bold=True)
            red_font = Font(color="FF0000", bold=True)
            try:
                diff_col_letter = get_column_letter(df_final_export.columns.get_loc('Diferencia') + 1)
                for row_idx in range(2, worksheet.max_row + 1):
                    cell_ref = f"{diff_col_letter}{row_idx}"
                    cell = worksheet[cell_ref]
                    if cell.value is not None:
                        try:
                            diff_value = int(cell.value)
                            if diff_value > 0: cell.font = blue_font
                            elif diff_value < 0: cell.font = red_font
                        except (ValueError, TypeError): pass
            except KeyError:
                print("API Resumen: Columna 'Diferencia' no encontrada para aplicar formato.")
            for i, col_name in enumerate(df_final_export.columns):
                 column_letter = get_column_letter(i + 1)
                 max_len = max(
                     df_final_export[col_name].astype(str).map(len).max(),
                     len(col_name)
                 ) + 2
                 worksheet.column_dimensions[column_letter].width = max_len
        output.seek(0)
        timestamp_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"resumen_por_grn_{timestamp_str}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error durante la generación del resumen por GRN Excel: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Error interno al generar el archivo de resumen por GRN"}), 500

# --- Ruta para servir el archivo HTML principal ---
@app.route('/Registro_inbound')
def registro_inbound_page():
    return render_template('inbound.html')

# --- Lógica de inicialización ---
if not os.path.exists(DATABASE_FOLDER):
    os.makedirs(DATABASE_FOLDER)
    print(f"FS: Carpeta de bases de datos creada: {DATABASE_FOLDER}")
templates_dir = os.path.join(APP_ROOT, 'templates')
if not os.path.exists(templates_dir):
    os.makedirs(templates_dir)
    print(f"FS: Carpeta de plantillas creada: {templates_dir}")
    placeholder_html_path = os.path.join(templates_dir, 'inbound.html')
    if not os.path.exists(placeholder_html_path):
        with open(placeholder_html_path, 'w', encoding='utf-8') as f:
            f.write("<h1>Placeholder para inbound.html</h1><p>Por favor, coloca tu archivo inbound.html aquí.</p>")
        print(f"FS: Archivo inbound.html de placeholder creado en {templates_dir}")
static_dir = os.path.join(APP_ROOT, 'static')
if not os.path.exists(static_dir):
    os.makedirs(static_dir)
    print(f"FS: Carpeta estática creada: {static_dir}")
    static_images_dir = os.path.join(static_dir, 'images')
    if not os.path.exists(static_images_dir):
       os.makedirs(static_images_dir)
       print(f"FS: Carpeta static/images creada.")

init_db() 

if not os.path.exists(ITEM_MASTER_CSV_PATH):
     print(f"FS ADVERTENCIA: Archivo maestro no encontrado. Creando ejemplo en: {ITEM_MASTER_CSV_PATH}")
     example_master_df = pd.DataFrame({
         'Item_Code': ['BG1234567890123', 'FT9876543210987', 'OTRO_ITEM_001'],
         'Item_Description': ['Maintenance Kit 1000h', 'Filtro de Aceite Modelo X', 'Repuesto Genérico Alfa'],
         'Weight_per_Unit': ['10 kg', '2 kg', '0.5 kg'],
         'Bin_1': ['RA25A', 'SB10C', 'ZC01X'],
         'Aditional_Bin_Location': ['RA25A', 'SB10C, SB11A', 'ZC01X']
     })
     example_master_df.to_csv(ITEM_MASTER_CSV_PATH, index=False, encoding='utf-8')
if not os.path.exists(GRN_CSV_FILE_PATH):
     print(f"FS ADVERTENCIA: Archivo GRN no encontrado. Creando ejemplo en: {GRN_CSV_FILE_PATH}")
     example_grn_df = pd.DataFrame({
         GRN_COLUMN_NAME_IN_CSV: ['21044', '21044', '21048', '21049'],
         'Item_Code': ['BG01499917', 'FT9876543210987', 'BG01499917', 'OTRO_ITEM_001'],
         'Quantity': [12, 5, 8, 20]
     })
     example_grn_df.to_csv(GRN_CSV_FILE_PATH, index=False, encoding='utf-8')

# No app.run() aquí
